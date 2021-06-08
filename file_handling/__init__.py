import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]
product_per_supplier = {}
total_value_per_supplier = {}
max_rows = product_list.max_row
print(max_rows)

# calculation for product per supplier
for row in range(2, max_rows + 1):
    supplier = product_list.cell(row, 4).value
    if supplier in product_per_supplier:
        product_per_supplier[supplier] = product_per_supplier.get(supplier) + 1
    else:
        product_per_supplier[supplier] = 1

print(f"Product per supplier : \n {product_per_supplier}")

for ele in product_per_supplier:
    print(f"{ele} {product_per_supplier[ele]}")

for row in range(2, max_rows + 1):
    supplier = product_list.cell(row, 4).value
    inventory = product_list.cell(row, 2).value
    price = product_list.cell(row, 3).value
    if supplier in total_value_per_supplier:
        total_value_per_supplier[supplier] = total_value_per_supplier[supplier] + (price * inventory)
    else:
        total_value_per_supplier[supplier] = price * inventory

    # add column in the file
    inventory_column = product_list.cell(row,5)
    inventory_column.value = total_value_per_supplier[supplier]

for ele in total_value_per_supplier:
    print(f"{ele} {total_value_per_supplier[ele]}")

inv_file.save("updated_inventory.xls")

